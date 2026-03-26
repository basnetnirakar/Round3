from __future__ import annotations

import zipfile
from pathlib import Path
from typing import Dict, List, Optional
import uuid
import os
import tempfile
import requests

import pandas as pd
import streamlit as st
from lxml import etree
from openpyxl import load_workbook
from io import BytesIO

# Optional: Pillow for better image resampling when resizing chromatograms
try:
    from PIL import Image
    PIL_AVAILABLE = True
except Exception:
    PIL_AVAILABLE = False

# Optional dependency: streamlit-aggrid enables clickable table rows
try:
    from st_aggrid import AgGrid, GridOptionsBuilder, GridUpdateMode
    AGGRID_AVAILABLE = True
except Exception:
    AGGRID_AVAILABLE = False

# ============================================================
# Antibody Data Manager for the exact Round3.xlsx layout
#
# Expected file layout:
#   project/
#     app.py
#     Round3.xlsx
#
# Workbook layout expected from the provided example:
#   Sheet name: Sheet1
#   Columns:
#     A  Antibody_name
#     B  Human EC50 (nM)
#     C  EC50_Image          (Excel in-cell image / rich value)
#     D  SPR_binding
#     E  ka(1/Ms)
#     F  kd(1/s)
#     G  KD
#     H  SPR-Mu              (Excel in-cell image / rich value)
#     I  Remarks
#
# Notes:
# - The images in columns C and H are stored as Excel rich-value cell images.
# - Normal pandas/openpyxl reading will often show '#VALUE!' in those cells.
# - This app reads the row data with pandas and separately extracts the actual
#   image bytes by parsing the workbook zip structure.
#
# Install:
#   pip install streamlit pandas openpyxl lxml
#
# Run:
#   streamlit run app.py
# ============================================================

WORKBOOK_LOCAL = Path("Round3_2.xlsx")
WORKBOOK_URL = os.environ.get("WORKBOOK_URL")  # set this in the hosting environment if using Drive

# Workbook resolution: prefer the newer Round3_2.xlsx but fall back to Round3.xlsx
SHEET_NAME = "Sheet1"
EXCEL_ROW_OFFSET = 2  # pandas row 0 corresponds to Excel row 2 because row 1 is headers

DISPLAY_COLUMNS = [
    "Antibody_name",
    "Human EC50 (nM)",
    "SPR_binding",
    "ka(1/Ms)",
    "kd(1/s)",
    "KD",
    "Remarks",
    # New SEC fields in updated workbook
    "SEC profile PSR",
    "SEC value",
]

st.set_page_config(page_title="Antibody Data Manager", layout="wide")


# -----------------------------
# Utility helpers
# -----------------------------
def format_scalar(value: object) -> str:
    if pd.isna(value):
        return ""
    return str(value)


def index_to_excel_col(idx: int) -> str:
    """Convert 0-based column index to Excel column letters (0 -> A)."""
    letters = ""
    i = idx + 1
    while i > 0:
        i, rem = divmod(i - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


@st.cache_data(show_spinner=False)
def read_main_table(workbook_path: str) -> pd.DataFrame:
    df = pd.read_excel(workbook_path, sheet_name=SHEET_NAME)
    df.columns = [str(c).strip() for c in df.columns]
    return df


@st.cache_data(show_spinner=False)
def read_workbook_info(workbook_path: str) -> Dict[str, object]:
    wb = load_workbook(workbook_path, data_only=False)
    ws = wb[SHEET_NAME]
    return {
        "sheet_names": wb.sheetnames,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
    }


@st.cache_data(show_spinner=False)
def extract_cell_images_from_xlsx(workbook_path: str) -> Dict[str, bytes]:
    """
    Return a mapping of Excel cell reference to image bytes.

    Example keys:
      C2, H2, C3, H3, ...

    This is tailored to the provided Round3.xlsx pattern, where images are
    stored as rich-value cell images and cells carry a vm attribute that maps
    to rId entries under xl/richData/_rels/richValueRel.xml.rels.
    """
    image_map: Dict[str, bytes] = {}

    with zipfile.ZipFile(workbook_path, "r") as z:
        names = set(z.namelist())
        rels_name = "xl/richData/_rels/richValueRel.xml.rels"
        sheet_xml_name = "xl/worksheets/sheet1.xml"

        if rels_name not in names or sheet_xml_name not in names:
            return image_map

        rel_root = etree.fromstring(z.read(rels_name))
        rels: Dict[str, str] = {}
        for rel in rel_root:
            rel_id = rel.get("Id")
            target = rel.get("Target")
            if rel_id and target:
                rels[rel_id] = target

        sheet_root = etree.fromstring(z.read(sheet_xml_name))
        ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

        for cell in sheet_root.xpath(".//main:c[@vm]", namespaces=ns):
            cell_ref = cell.get("r")
            vm = cell.get("vm")
            if not cell_ref or not vm:
                continue

            rel_id = f"rId{vm}"
            target = rels.get(rel_id)
            if not target:
                continue

            media_path = "xl/" + target.replace("../", "")
            if media_path not in names:
                continue

            image_map[cell_ref] = z.read(media_path)

    return image_map


@st.cache_data(show_spinner=False)
def build_app_dataframe(workbook_path: str) -> pd.DataFrame:
    df = read_main_table(workbook_path).copy()
    image_map = extract_cell_images_from_xlsx(workbook_path)

    excel_rows: List[int] = []

    # detect image-like columns in two ways:
    # 1) header contains 'image' or 'chromatogram'
    # 2) any cell in that column has a matching extracted image in image_map
    image_columns: List[str] = []
    for idx, c in enumerate(df.columns):
        col_letter = index_to_excel_col(idx)
        header_hit = "image" in str(c).lower() or "chromatogram" in str(c).lower()
        has_image_in_map = any(k.startswith(col_letter) for k in image_map.keys())
        if header_hit or has_image_in_map:
            image_columns.append(c)

    # prepare per-column lists for image bytes
    image_bytes_map: Dict[str, List[Optional[bytes]]] = {c: [] for c in image_columns}

    for i in range(len(df)):
        excel_row = i + EXCEL_ROW_OFFSET
        excel_rows.append(excel_row)
        for col in image_columns:
            col_idx = int(df.columns.get_loc(col))
            col_letter = index_to_excel_col(col_idx)
            cell_ref = f"{col_letter}{excel_row}"
            image_bytes_map[col].append(image_map.get(cell_ref))

    df["excel_row"] = excel_rows

    # attach image bytes and boolean flags for each detected image column
    for col in image_columns:
        bytes_col = f"{col}_bytes"
        has_col = f"has_{col}_image"
        df[bytes_col] = image_bytes_map[col]
        df[has_col] = df[bytes_col].apply(lambda x: x is not None)

    return df


# -----------------------------
# App UI
# -----------------------------
st.title("Antibody Data Manager")
st.caption("Interactive browser for the exact Round3.xlsx layout, including in-cell EC50 and SPR images.")

def ensure_workbook_local():
    if WORKBOOK_LOCAL.exists():
        return WORKBOOK_LOCAL
    if WORKBOOK_URL:
        # download to a temp file (persist to repo folder so subsequent reads work)
        r = requests.get(WORKBOOK_URL, stream=True, timeout=30)
        r.raise_for_status()
        with open(WORKBOOK_LOCAL, "wb") as f:
            for chunk in r.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
        return WORKBOOK_LOCAL
    raise FileNotFoundError(f"Workbook not found locally and WORKBOOK_URL not set: {WORKBOOK_LOCAL}")

if not WORKBOOK_LOCAL.exists():
    st.error(f"Workbook not found: {WORKBOOK_LOCAL.resolve()}")
    st.info("Place Round3.xlsx in the same folder as app.py, then rerun the app.")
    st.stop()

try:
    app_df = build_app_dataframe(str(WORKBOOK_LOCAL))
    workbook_info = read_workbook_info(str(WORKBOOK_LOCAL))
except Exception as e:
    st.error(f"Failed to load workbook: {e}")
    st.stop()

# Sidebar filters
st.sidebar.header("Filters")
search_text = st.sidebar.text_input("Search antibody or remarks")
only_with_ec50_image = st.sidebar.checkbox("Only rows with EC50 image")
only_with_spr_image = st.sidebar.checkbox("Only rows with SPR image")
only_spr_binding_positive = st.sidebar.checkbox("Only SPR_binding = 1")

filtered_df = app_df.copy()

if search_text:
    q = search_text.strip()
    mask = (
        filtered_df["Antibody_name"].astype(str).str.contains(q, case=False, na=False)
        | filtered_df["Remarks"].astype(str).str.contains(q, case=False, na=False)
    )
    filtered_df = filtered_df[mask]

if only_with_ec50_image:
    filtered_df = filtered_df[filtered_df["has_ec50_image"]]

if only_with_spr_image:
    filtered_df = filtered_df[filtered_df["has_spr_image"]]

if only_spr_binding_positive:
    filtered_df = filtered_df[filtered_df["SPR_binding"].astype(str) == "1"]

st.subheader("Antibody table")

# Prepare display table: only include DISPLAY_COLUMNS that exist in the DataFrame
display_present = [c for c in DISPLAY_COLUMNS if c in filtered_df.columns]

# Collect any 'has_*_image' boolean flags that were added dynamically
has_flags = [c for c in filtered_df.columns if c.startswith("has_") and c.endswith("_image")]

# Build the show_table from available columns (don't assume optional SEC columns exist)
show_table = filtered_df[display_present + has_flags].copy()

# Rename has_* flags to friendly labels (e.g. has_ec50_image -> 'Ec50 image')
rename_map = {}
for flag in has_flags:
    # derive base name between 'has_' and '_image'
    base = flag[len("has_") : -len("_image")] if flag.endswith("_image") else flag
    label = f"{base.replace('_', ' ').strip().title()} image"
    rename_map[flag] = label

show_table = show_table.rename(columns=rename_map)

# Build ag_table from the columns that actually exist to avoid KeyError
ag_columns = display_present.copy()
for flag in has_flags:
    if flag not in ag_columns:
        ag_columns.append(flag)
if "excel_row" in filtered_df.columns and "excel_row" not in ag_columns:
    ag_columns.append("excel_row")

ag_table = filtered_df[ag_columns].copy()
ag_table = ag_table.rename(columns=rename_map)

if "selected_excel_row" not in st.session_state and not filtered_df.empty:
    st.session_state.selected_excel_row = int(filtered_df.iloc[0]["excel_row"])

if AGGRID_AVAILABLE:
    gb = GridOptionsBuilder.from_dataframe(ag_table)
    gb.configure_selection(selection_mode="single", use_checkbox=False)
    grid_options = gb.build()
    grid_response = AgGrid(
        ag_table,
        gridOptions=grid_options,
        height=300,
        fit_columns_on_grid_load=True,
        update_mode=GridUpdateMode.SELECTION_CHANGED,
    )

    def _normalize_selected_rows(obj):
        """Return a list of record dicts for various AgGrid response shapes."""
        if obj is None:
            return []
        # pandas DataFrame -> records
        try:
            import pandas as _pd

            if isinstance(obj, _pd.DataFrame):
                return obj.to_dict("records")
        except Exception:
            pass

        # dict -> single-item list
        if isinstance(obj, dict):
            return [obj]

        # list-like (including tuples)
        if isinstance(obj, (list, tuple)):
            return list(obj)

        # some responses may be numpy arrays or other sequence types
        try:
            return list(obj)
        except Exception:
            return []

    selected_rows_raw = grid_response.get("selected_rows", [])
    selected_rows = _normalize_selected_rows(selected_rows_raw)

    if len(selected_rows) > 0:
        try:
            sel_row = selected_rows[0]
            # extract excel_row robustly
            sel_excel = None
            if isinstance(sel_row, dict):
                sel_excel = sel_row.get("excel_row")
                if sel_excel is None:
                    for k, v in sel_row.items():
                        if str(k).lower().replace(" ", "_") == "excel_row":
                            sel_excel = v
                            break
            # coerce numeric-like values
            if sel_excel is not None:
                try:
                    st.session_state.selected_excel_row = int(sel_excel)
                except Exception:
                    st.error("Selected row excel_row value could not be converted to int")
        except Exception as e:
            # show useful debug info in the app but don't crash
            st.error(f"Selection handling error: {e}")
            st.write("AgGrid selection response (raw):", selected_rows_raw)
            st.write("AgGrid selection response (normalized):", selected_rows)
    # AgGrid is already rendered above; do not render the non-interactive table.
    pass
else:
    st.info("To enable clickable row selection install `streamlit-aggrid` (conda install -c conda-forge streamlit-aggrid).\nInteractive selection will be available after restarting the app.")

if filtered_df.empty:
    st.info("No rows match the selected filters.")
    st.stop()

# Ensure selected_excel_row exists and is valid for the current filtered_df
if not filtered_df.empty:
    if "selected_excel_row" not in st.session_state:
        st.session_state.selected_excel_row = int(filtered_df.iloc[0]["excel_row"])

# per-session unique token for widget keys to avoid duplicate-key collisions
if "_ui_key_token" not in st.session_state:
    st.session_state["_ui_key_token"] = uuid.uuid4().hex

    if st.session_state.selected_excel_row not in filtered_df["excel_row"].values:
        st.session_state.selected_excel_row = int(filtered_df.iloc[0]["excel_row"])

selected_row = filtered_df[filtered_df["excel_row"] == st.session_state.selected_excel_row].iloc[0]

st.markdown(f"## Details: {selected_row['Antibody_name']}")
left, right = st.columns([1, 1])

with left:
    # Summary moved below the Images tab per UI change request.
    st.write("")

with right:
    # Workbook info moved below Summary in the Images tab.
    st.write("")

image_tab, raw_tab, gallery_tab = st.tabs(["Images", "Raw row", "Gallery"])

with image_tab:
    # Build image column list from dynamically-added has_* flags so we only
    # show columns that actually had extracted images attached.
    has_flags_local = [c for c in app_df.columns if c.startswith("has_") and c.endswith("_image")]
    image_columns = [f[len("has_") : -len("_image")] for f in has_flags_local]

    # Reliable Prev / Next buttons so users can navigate without relying on
    # keyboard capture or AgGrid focus. These update session state and rerun.
    nav_col1, nav_col2 = st.columns([1, 1])
    with nav_col1:
        if st.button("← Prev"):
            try:
                excel_rows_list = list(filtered_df["excel_row"].astype(int))
                cur = int(st.session_state.selected_excel_row)
                pos = excel_rows_list.index(cur) if cur in excel_rows_list else 0
                new_pos = max(pos - 1, 0)
                st.session_state.selected_excel_row = int(excel_rows_list[new_pos])
                st.experimental_rerun()
            except Exception:
                pass
    with nav_col2:
        if st.button("Next →"):
            try:
                excel_rows_list = list(filtered_df["excel_row"].astype(int))
                cur = int(st.session_state.selected_excel_row)
                pos = excel_rows_list.index(cur) if cur in excel_rows_list else 0
                new_pos = min(pos + 1, len(excel_rows_list) - 1)
                st.session_state.selected_excel_row = int(excel_rows_list[new_pos])
                st.experimental_rerun()
            except Exception:
                pass

    # Show the raw row values above images for quick context (mirrors the Raw tab)
    st.markdown("### Raw row (quick view)")
    raw_columns = [
        c
        for c in app_df.columns
        if not str(c).lower().endswith("_bytes") and not str(c).lower().startswith("has_")
    ]
    if "excel_row" not in raw_columns:
        raw_columns.append("excel_row")

    raw_display = pd.DataFrame([selected_row[raw_columns].to_dict()])
    st.dataframe(raw_display, use_container_width=True, hide_index=True)
    st.caption(
        "Note: EC50_Image and SPR-Mu often show '#VALUE!' in raw Excel reads because the true images are stored separately as rich-value cell images."
    )

    if not image_columns:
        st.info("No image-like columns were detected in the workbook.")
    else:
        # Prepare list of (col_name, bytes) for the selected row
        images_to_show = []
        for col in image_columns:
            bytes_col = f"{col}_bytes"
            img_bytes = selected_row.get(bytes_col)
            if img_bytes is not None:
                images_to_show.append((col, img_bytes))

        # Limit to max 4 images for a 2x2 grid
        images_to_show = images_to_show[:4]

        # Create rows of 2 images each
        rows = [images_to_show[i : i + 2] for i in range(0, len(images_to_show), 2)]

        # Render grid: 2 columns per row
        for row_imgs in rows:
            cols = st.columns(2)
            for idx, (col_name, img_bytes) in enumerate(row_imgs):
                with cols[idx]:
                    st.markdown(f"**{col_name}**")
                    try:
                        if PIL_AVAILABLE:
                            img = Image.open(BytesIO(img_bytes))
                            if "chromatogram" in col_name.lower():
                                target_width = 200
                            else:
                                target_width = 600

                            if img.width > target_width:
                                ratio = target_width / float(img.width)
                                target_height = int(img.height * ratio)
                                img = img.resize((target_width, target_height), Image.LANCZOS)
                            st.image(img, width=target_width)
                        else:
                            # Streamlit width fallback
                            if "chromatogram" in col_name.lower():
                                st.image(img_bytes, width=200)
                            else:
                                st.image(img_bytes, width=600)
                    except Exception:
                        # fallback width based on column type
                        if "chromatogram" in col_name.lower():
                            st.image(img_bytes, width=200)
                        else:
                            st.image(img_bytes, width=600)

        # If less than 4 images, optionally show placeholders for visual balance
        # (we simply leave empty space in the grid)

    # Summary (moved below Images)
    st.markdown("### Summary")
    # Use the display_present columns (intersection of DISPLAY_COLUMNS and actual columns)
    try:
        summary_fields = display_present.copy()
    except NameError:
        # fallback: use DISPLAY_COLUMNS that exist in app_df
        summary_fields = [c for c in DISPLAY_COLUMNS if c in app_df.columns]

    # Always include excel_row at the end
    if "excel_row" not in summary_fields:
        summary_fields.append("excel_row")

    summary_rows = []
    for f in summary_fields:
        summary_rows.append({"Field": f, "Value": format_scalar(selected_row.get(f))})

    summary_df = pd.DataFrame(summary_rows)
    st.table(summary_df)

    st.download_button(
        label="Download Round3.xlsx",
        data=WORKBOOK_LOCAL.read_bytes(),
        file_name=WORKBOOK_LOCAL.name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    # Workbook info (moved below Summary)
    st.markdown("### Workbook info")
    st.write(f"**File:** {WORKBOOK_LOCAL.name}")
    st.write(f"**Sheet names:** {', '.join(workbook_info['sheet_names'])}")
    st.write(f"**Rows:** {workbook_info['max_row']}")
    st.write(f"**Columns:** {workbook_info['max_column']}")
    # dynamic image mapping
    image_columns_ui = [f[len("has_") : -len("_image")] for f in app_df.columns if f.startswith("has_") and f.endswith("_image")]
    if image_columns_ui:
        mapping = ", ".join(
            f"{index_to_excel_col(list(app_df.columns).index(c))} = {c}" for c in image_columns_ui
        )
        st.write(f"**Image mapping:** {mapping}")
    else:
        st.write("**Image mapping:** (none detected)")

with raw_tab:
    st.markdown("### Raw row values")
    # Show raw workbook columns (exclude internal _bytes and has_ columns)
    raw_columns = [
        c
        for c in app_df.columns
        if not str(c).lower().endswith("_bytes") and not str(c).lower().startswith("has_")
    ]
    # ensure excel_row is included
    if "excel_row" not in raw_columns:
        raw_columns.append("excel_row")

    raw_display = pd.DataFrame([selected_row[raw_columns].to_dict()])
    st.dataframe(raw_display, use_container_width=True, hide_index=True)
    st.caption("Note: EC50_Image and SPR-Mu often show '#VALUE!' in raw Excel reads because the true images are stored separately as rich-value cell images.")

with gallery_tab:
    st.markdown("### All extracted images for filtered rows")
    # dynamic gallery based on detected image-like columns
    image_columns_gallery = [f[len("has_") : -len("_image")] for f in app_df.columns if f.startswith("has_") and f.endswith("_image")]

    if not image_columns_gallery:
        st.info("No extracted images available for gallery.")
    else:
        gallery_choice = st.radio("Gallery type", image_columns_gallery, horizontal=True)

        for _, row in filtered_df.iterrows():
            bytes_col = f"{gallery_choice}_bytes"
            image_bytes = row.get(bytes_col)
            if image_bytes is None:
                continue
            st.markdown(f"**{row['Antibody_name']}**")
            # Small thumbnail-style images in gallery; let Streamlit pick a reasonable width
            if PIL_AVAILABLE:
                try:
                    img = Image.open(BytesIO(image_bytes))
                    # scale down large images for gallery
                    max_w = 600
                    if img.width > max_w:
                        ratio = max_w / float(img.width)
                        img = img.resize((max_w, int(img.height * ratio)), Image.LANCZOS)
                    st.image(img, width=max_w)
                except Exception:
                    st.image(image_bytes, width=600)
            else:
                st.image(image_bytes, width=600)

st.markdown("---")
st.markdown("### Notes for this workbook")
st.write(
    "This app is tailored to the provided Round3.xlsx structure. "
    "It assumes Sheet1 contains one antibody per row, with in-cell images stored as rich-value cell images. "
    "Detected image columns (EC50, SPR, SEC chromatogram) are rendered in the Images tab and in the Gallery."
)
